import numpy as np
from openpyxl import load_workbook
import pickle
if __name__ == '__main__':
    wb = load_workbook("./data/SWaT/SWaT_train.xlsx")
    ws = wb.worksheets[0]
    row = ws.max_row
    col = ws.max_column
    data_og = []
    name = []
    for i in range(2, col):#名字
        name.append(ws.cell(1, i).value.replace(' ', ''))
    for i in range(2, row + 1):#数据
        data_og.append([])
        for j in range(2, col):
            data_og[-1].append(float(ws.cell(i, j).value))
    values = np.array(data_og)#values存所有的值
    dataDict = {}
    dataDict['name'] = name
    dataDict['values'] = values
    fw = open('./file/data_normal_v1_pickle.txt', 'wb')
    pickle.dump(dataDict, fw)
    fw.close()

    wb = load_workbook("./data/SWaT/SWaT_test.xlsx")
    ws = wb.worksheets[0]
    row = ws.max_row
    col = ws.max_column
    data_og = []
    name = []
    label = []
    for i in range(2, col):
        name.append(ws.cell(1, i).value.replace(' ', ''))
    for i in range(2, row + 1):
        label.append(ws.cell(i, col).value)
    for i in range(2, row + 1):
        data_og.append([])
        for j in range(2, col):
            data_og[-1].append(float(ws.cell(i, j).value))
    values = np.array(data_og)
    dataDict = {}
    dataDict['name'] = name
    dataDict['values'] = values
    dataDict['label'] = label
    print(values.shape)
    print(len(name), len(label))
    fw = open('./file/data_attack_v0_pickle.txt', 'wb')
    pickle.dump(dataDict, fw)
    fw.close()